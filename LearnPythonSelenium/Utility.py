import openpyxl
from selenium import webdriver
class ExcelUtility:
    def __init__(self,excelFile,sheetName,driver):
        self.excelFile = excelFile
        self.sheetName = sheetName
        self.driver = driver
        self.rowCount=0
        self.columnCount=0
    def setRowColumn(self):
        workBook = openpyxl.load_workbook(self.excelFile)
        sheet = workBook[self.sheetName]
        self.rowCount = sheet.max_row
        self.columnCount = sheet.max_column
    def readData(self,row,column):
        workBook = openpyxl.load_workbook(self.excelFile)
        sheet = workBook[self.sheetName]
        values = sheet.cell(row,column).value
        return values
